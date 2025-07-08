#!/usr/bin/env python3
"""
Script d'analyse de configuration nxlog
Analyse un ou plusieurs fichiers de configuration nxlog et affiche les paramètres sous forme de tableau
Inclut la cartographie des flux et des routes
"""

import re
import sys
import argparse
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set
import os
import glob

try:
    from tabulate import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False
    print("Module 'tabulate' requis. Installez-le avec: python3 -m pip install tabulate")
    print("Ou utilisez le format CSV/JSON en attendant: --format csv ou --format json")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class FlowMapper:
    """Classe pour analyser et cartographier les flux NXLog"""
    
    def __init__(self):
        self.routes = {}  # {route_name: {path: str, priority: int, condition: str}}
        self.sections = {}  # {section_name: {type: str, module: str, details: dict}}
        self.flows = []  # Liste des flux analysés
        self.connections = {}  # Graphe des connexions
        self.sections_info = {}
        self.unconnected_sections = set()
        self.graphviz_content = ""
    
    def add_route(self, route_name: str, path: str, priority: int = None, condition: str = None):
        """Ajoute une route à la cartographie"""
        self.routes[route_name] = {
            'path': path,
            'priority': priority,
            'condition': condition,
            'inputs': [],
            'outputs': [],
            'processors': []
        }
        
        # Parser le chemin de la route
        self._parse_route_path(route_name, path)
    
    def _parse_route_path(self, route_name: str, path: str):
        """Parse le chemin d'une route pour identifier les connexions"""
        # Format typique: "input1, input2 => processor1 => output1, output2"
        # ou: "input1 => output1"
        
        if '=>' in path:
            parts = [part.strip() for part in path.split('=>')]
            
            # Premier élément = inputs
            if parts[0]:
                inputs = [inp.strip() for inp in parts[0].split(',')]
                self.routes[route_name]['inputs'] = inputs
            
            # Dernier élément = outputs
            if len(parts) > 1 and parts[-1]:
                outputs = [out.strip() for out in parts[-1].split(',')]
                self.routes[route_name]['outputs'] = outputs
            
            # Éléments intermédiaires = processors
            if len(parts) > 2:
                for i in range(1, len(parts) - 1):
                    if parts[i]:
                        processors = [proc.strip() for proc in parts[i].split(',')]
                        self.routes[route_name]['processors'].extend(processors)
    
    def add_section(self, section_name: str, section_type: str, module: str = None, details: dict = None):
        """Ajoute une section à la cartographie"""
        self.sections[section_name] = {
            'type': section_type,
            'module': module,
            'details': details or {},
            'connected_routes': []
        }
        
        self.sections_info[section_name] = {
            'type': section_type,
            'module': module or '',
            'details': details or {}
        }
    
    def analyze_flows(self):
        """Analyse les flux entre les sections"""
        self.flows = []
        
        for route_name, route_info in self.routes.items():
            # Créer les flux pour cette route
            inputs = route_info['inputs']
            processors = route_info['processors']
            outputs = route_info['outputs']
            
            # Marquer les sections comme connectées à cette route
            for section_name in inputs + processors + outputs:
                if section_name in self.sections:
                    self.sections[section_name]['connected_routes'].append(route_name)
            
            # Créer les connexions directes
            all_components = inputs + processors + outputs
            
            for i in range(len(all_components) - 1):
                source = all_components[i]
                target = all_components[i + 1]
                
                flow = {
                    'route': route_name,
                    'source': source,
                    'target': target,
                    'destination': target,
                    'source_type': self.sections.get(source, {}).get('type', 'Unknown'),
                    'target_type': self.sections.get(target, {}).get('type', 'Unknown'),
                    'source_module': self.sections.get(source, {}).get('module', 'Unknown'),
                    'target_module': self.sections.get(target, {}).get('module', 'Unknown'),
                    'priority': route_info.get('priority') or 'N/A',
                    'condition': route_info.get('condition') or 'N/A'
                }
                
                self.flows.append(flow)
        
        # Identifier les sections non connectées
        connected_sections = set()
        for flow in self.flows:
            connected_sections.add(flow['source'])
            connected_sections.add(flow['target'])
        
        self.unconnected_sections = set(self.sections_info.keys()) - connected_sections
        
        # Générer le contenu Graphviz
        self._generate_graphviz()
        
        return {
            'flows': self.flows,
            'sections': dict(self.sections_info),
            'unconnected': list(self.unconnected_sections),
            'stats': self._get_flow_stats(),
            'graphviz': self.graphviz_content
        }
    
    def _generate_graphviz(self):
        """Génère le contenu Graphviz pour la visualisation"""
        dot_content = []
        dot_content.append('digraph NXLogFlow {')
        dot_content.append('    rankdir=LR;')
        dot_content.append('    node [shape=box, style=filled];')
        dot_content.append('    edge [fontsize=10];')
        dot_content.append('')
        
        # Définir les couleurs par type de section
        colors = {
            'Input': '#E8F5E8',      # Vert clair
            'Output': '#FFE8E8',     # Rouge clair
            'Processor': '#E8E8FF',  # Bleu clair
            'Extension': '#FFFFE8',  # Jaune clair
            'Route': '#F0F0F0'       # Gris clair
        }
        
        # Ajouter les nœuds (sections)
        dot_content.append('    // Sections')
        for section_name, info in self.sections_info.items():
            section_type = info['type']
            module = info['module']
            color = colors.get(section_type, '#FFFFFF')
            
            # Créer le label avec type et module
            label = f"{section_name}\\n({section_type})"
            if module:
                label += f"\\n{module}"
            
            # Ajouter des attributs spéciaux pour les sections non connectées
            if section_name in self.unconnected_sections:
                dot_content.append(f'    "{section_name}" [label="{label}", fillcolor="{color}", style="filled,dashed"];')
            else:
                dot_content.append(f'    "{section_name}" [label="{label}", fillcolor="{color}"];')
        
        dot_content.append('')
        
        # Ajouter les connexions (flux)
        dot_content.append('    // Flux de données')
        route_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD']
        route_color_map = {}
        color_index = 0
        
        for flow in self.flows:
            route = flow['route']
            source = flow['source']
            destination = flow['destination']
            priority = flow['priority']
            condition = flow['condition']
            
            # Assigner une couleur par route
            if route not in route_color_map:
                route_color_map[route] = route_colors[color_index % len(route_colors)]
                color_index += 1
            
            edge_color = route_color_map[route]
            
            # Créer le label de l'arête
            edge_label = f"Route: {route}"
            if priority != 'N/A':
                edge_label += f"\\nPriorité: {priority}"
            if condition != 'N/A':
                # Raccourcir les conditions trop longues
                short_condition = condition[:30] + "..." if len(condition) > 30 else condition
                edge_label += f"\\nCondition: {short_condition}"
            
            dot_content.append(f'    "{source}" -> "{destination}" [label="{edge_label}", color="{edge_color}", fontcolor="{edge_color}"];')
        
        dot_content.append('')
        
        # Ajouter une légende
        dot_content.append('    // Légende')
        dot_content.append('    subgraph cluster_legend {')
        dot_content.append('        label="Légende";')
        dot_content.append('        style=filled;')
        dot_content.append('        fillcolor="#F5F5F5";')
        dot_content.append('        fontsize=12;')
        dot_content.append('')
        
        for section_type, color in colors.items():
            if any(info['type'] == section_type for info in self.sections_info.values()):
                dot_content.append(f'        "legend_{section_type}" [label="{section_type}", fillcolor="{color}", shape=box];')
        
        dot_content.append('        "legend_disconnected" [label="Non connecté", fillcolor="#FFFFFF", style="filled,dashed", shape=box];')
        dot_content.append('    }')
        dot_content.append('')
        
        dot_content.append('}')
        
        self.graphviz_content = '\n'.join(dot_content)
    
    def _parse_path_expression(self, path_expr):
        """Parse une expression de chemin comme 'input1, input2 => processor1 => output1, output2'"""
        # Nettoyer l'expression
        path_expr = path_expr.strip()
        
        # Diviser par '=>'
        segments = [seg.strip() for seg in path_expr.split('=>')]
        
        flows = []
        for i in range(len(segments) - 1):
            sources = [s.strip() for s in segments[i].split(',')]
            destinations = [d.strip() for d in segments[i + 1].split(',')]
            
            # Créer tous les flux possibles entre sources et destinations
            for source in sources:
                for dest in destinations:
                    if source and dest:
                        flows.append((source.strip(), dest.strip()))
        
        return flows
    
    def _get_flow_stats(self):
        """Calcule les statistiques des flux"""
        stats = {
            'total_routes': len(self.routes),
            'total_sections': len(self.sections_info),
            'total_flows': len(self.flows),
            'unconnected_sections': len(self.unconnected_sections)
        }
        
        # Compter par type de section
        type_counts = {}
        for info in self.sections_info.values():
            section_type = info['type']
            type_counts[section_type] = type_counts.get(section_type, 0) + 1
        
        stats.update(type_counts)
        return stats
    
    def get_flow_summary(self):
        """Retourne un résumé des flux"""
        summary = {
            'total_routes': len(self.routes),
            'total_sections': len(self.sections),
            'total_flows': len(self.flows),
            'input_sections': len([s for s in self.sections.values() if s['type'] == 'Input']),
            'output_sections': len([s for s in self.sections.values() if s['type'] == 'Output']),
            'processor_sections': len([s for s in self.sections.values() if s['type'] == 'Processor']),
            'extension_sections': len([s for s in self.sections.values() if s['type'] == 'Extension']),
            'unconnected_sections': len([s for s in self.sections.values() if not s['connected_routes']])
        }
        return summary
    
    def get_flow_table_data(self):
        """Retourne les données des flux pour affichage en tableau"""
        table_data = []
        
        for flow in self.flows:
            table_data.append([
                flow['route'],
                flow['source'],
                flow['source_type'],
                flow['source_module'],
                '→',
                flow['target'],
                flow['target_type'],
                flow['target_module'],
                flow['priority'] or 'N/A',
                flow['condition'] or 'N/A'
            ])
        
        return table_data
    
    def get_sections_table_data(self):
        """Retourne les données des sections pour affichage en tableau"""
        table_data = []
        
        for section_name, section_info in self.sections.items():
            connected_routes = ', '.join(section_info['connected_routes']) if section_info['connected_routes'] else 'Non connecté'
            
            table_data.append([
                section_name,
                section_info['type'],
                section_info['module'] or 'N/A',
                connected_routes,
                len(section_info['connected_routes'])
            ])
        
        return table_data

class NXLogConfigAnalyzer:
    """Analyseur de fichier de configuration nxlog"""
    
    def __init__(self):
        self.config_data = []
        self.files_data = {}  # Dictionnaire pour stocker les données par fichier
        self.flow_mappers = {}  # Dictionnaire pour stocker les cartographies par fichier
        self.parameter_descriptions = {
            # Paramètres généraux
            'Module': 'Type de module utilisé (im_file, om_file, etc.)',
            'File': 'Chemin du fichier de log',
            'InputType': 'Type d\'entrée des données',
            'OutputType': 'Type de sortie des données',
            'Host': 'Adresse IP ou nom d\'hôte',
            'Port': 'Port de connexion',
            'Protocol': 'Protocole de communication (TCP/UDP)',
            'Facility': 'Facilité syslog',
            'Severity': 'Niveau de sévérité',
            'SourceName': 'Nom de la source',
            'Exec': 'Commande ou script à exécuter (peut être multi-lignes)',
            'Schedule': 'Planification d\'exécution',
            'SavePos': 'Sauvegarde de la position de lecture',
            'ReadFromLast': 'Lecture depuis la dernière position',
            'PollInterval': 'Intervalle de polling en secondes',
            'DirCheckInterval': 'Intervalle de vérification des répertoires',
            'ActiveFiles': 'Nombre maximum de fichiers actifs',
            'CloseWhenIdle': 'Fermer les fichiers inactifs',
            'Format': 'Format des données',
            'CSVDelimiter': 'Délimiteur CSV',
            'CSVQuoteChar': 'Caractère de guillemet CSV',
            'Fields': 'Champs définis',
            'Types': 'Types de données',
            'Reconnect': 'Reconnexion automatique',
            'ConnectionTimeout': 'Timeout de connexion',
            'FlushInterval': 'Intervalle de flush',
            'SyncInterval': 'Intervalle de synchronisation',
            'CreateDir': 'Créer le répertoire s\'il n\'existe pas',
            'Truncate': 'Tronquer le fichier',
            'Sync': 'Synchronisation forcée',
            'BufferSize': 'Taille du buffer',
            'LogLevel': 'Niveau de log',
            'LogFile': 'Fichier de log',
            'LogToConsole': 'Affichage des logs sur la console',
            'SpoolDir': 'Répertoire de spool',
            'CacheDir': 'Répertoire de cache',
            'PidFile': 'Fichier PID',
            'RootDir': 'Répertoire racine',
            'User': 'Utilisateur d\'exécution',
            'Group': 'Groupe d\'exécution',
            'ModuleDir': 'Répertoire des modules',
            'IncludeDir': 'Répertoire d\'inclusion',
            'SharedDir': 'Répertoire partagé',
            'DefaultCharset': 'Jeu de caractères par défaut',
            'NoCache': 'Désactiver le cache',
            'Threads': 'Nombre de threads',
            'MaxEvents': 'Nombre maximum d\'événements',
            'MaxQueueSize': 'Taille maximale de la queue',
            'Route': 'Route de traitement',
            'Path': 'Chemin de traitement',
            'Priority': 'Priorité',
            'Condition': 'Condition d\'exécution',
            'Drop': 'Abandonner l\'événement',
            'Transform': 'Transformation des données',
            'Regex': 'Expression régulière',
            'Pattern': 'Motif de recherche',
            'Replacement': 'Chaîne de remplacement',
            'Global': 'Remplacement global',
            'CaseSensitive': 'Sensible à la casse',
            'Multiline': 'Multi-lignes',
            'DotAll': 'Point correspond à tout',
            'Extended': 'Mode étendu',
            'Ungreedy': 'Mode non-gourmand',
            'Key': 'Clé de configuration',
            'Value': 'Valeur de configuration',
            'SSL': 'Utiliser SSL/TLS',
            'CertFile': 'Fichier de certificat',
            'CertKeyFile': 'Fichier de clé privée',
            'CAFile': 'Fichier CA',
            'CRLFile': 'Fichier CRL',
            'AllowUntrusted': 'Autoriser les certificats non fiables',
            'Compression': 'Compression des données',
            'Gzip': 'Compression gzip',
            'Bzip2': 'Compression bzip2',
            'XZ': 'Compression xz',
            'Password': 'Mot de passe',
            'Username': 'Nom d\'utilisateur',
            'Domain': 'Domaine',
            'Workstation': 'Nom de la station de travail',
            'HTTPSProxy': 'Proxy HTTPS',
            'HTTPProxy': 'Proxy HTTP',
            'ProxyUser': 'Utilisateur proxy',
            'ProxyPass': 'Mot de passe proxy',
            'ContentType': 'Type de contenu',
            'Headers': 'En-têtes HTTP',
            'URL': 'URL de destination',
            'HTTPSCertFile': 'Fichier de certificat HTTPS',
            'HTTPSCertKeyFile': 'Fichier de clé privée HTTPS',
            'HTTPSCAFile': 'Fichier CA HTTPS',
            'HTTPSCRLFile': 'Fichier CRL HTTPS',
            'HTTPSAllowUntrusted': 'Autoriser les certificats HTTPS non fiables',
            'JSONDateFormat': 'Format de date JSON',
            'JSONDateField': 'Champ de date JSON',
            'JSONTimeFormat': 'Format d\'heure JSON',
            'JSONTimeField': 'Champ d\'heure JSON',
            'XMLDateFormat': 'Format de date XML',
            'XMLDateField': 'Champ de date XML',
            'XMLTimeFormat': 'Format d\'heure XML',
            'XMLTimeField': 'Champ d\'heure XML',
            'Delimiter': 'Délimiteur',
            'QuoteChar': 'Caractère de guillemet',
            'EscapeChar': 'Caractère d\'échappement',
            'EscapeControl': 'Échapper les caractères de contrôle',
            'UndefValue': 'Valeur indéfinie',
            'HeaderLine': 'Ligne d\'en-tête',
            'FieldTypes': 'Types de champs',
            'FieldNames': 'Noms de champs'
        }
    
    def parse_config_file(self, config_file: str) -> Tuple[List[Dict], FlowMapper]:
        """Parse le fichier de configuration nxlog et retourne les données + cartographie des flux"""
        try:
            with open(config_file, 'r', encoding='utf-8') as f:
                content = f.read()
        except FileNotFoundError:
            print(f"Erreur: Fichier '{config_file}' non trouvé")
            return [], FlowMapper()
        except Exception as e:
            print(f"Erreur lors de la lecture du fichier {config_file}: {e}")
            return [], FlowMapper()
        
        # Supprimer les commentaires
        content = re.sub(r'#.*$', '', content, flags=re.MULTILINE)
        
        current_section = None
        current_section_name = None
        lines = content.split('\n')
        i = 0
        file_data = []
        flow_mapper = FlowMapper()
        
        # Variables pour stocker les informations de section
        section_module = None
        section_details = {}
        
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
                
            # Début de section
            section_match = re.match(r'<(\w+)\s+([^>]+)>', line, re.IGNORECASE)
            if section_match:
                current_section = section_match.group(1)
                current_section_name = section_match.group(2)
                section_module = None
                section_details = {}
                i += 1
                continue
            
            # Fin de section
            if re.match(r'</\w+>', line, re.IGNORECASE):
                # Ajouter la section au flow mapper
                if current_section and current_section_name:
                    flow_mapper.add_section(
                        current_section_name,
                        current_section,
                        section_module,
                        section_details.copy()
                    )
                
                current_section = None
                current_section_name = None
                section_module = None
                section_details = {}
                i += 1
                continue
            
            # Vérifier si c'est le début d'un bloc <Exec>
            if re.match(r'<Exec>', line, re.IGNORECASE):
                exec_lines = []
                i += 1  # Passer à la ligne suivante
                
                # Collecter toutes les lignes jusqu'à </Exec>
                while i < len(lines):
                    exec_line = lines[i].strip()
                    if re.match(r'</Exec>', exec_line, re.IGNORECASE):
                        break
                    if exec_line:  # Ignorer les lignes vides
                        exec_lines.append(exec_line)
                    i += 1
                
                # Ajouter le bloc Exec comme un seul paramètre
                if exec_lines and current_section:
                    param_value = '\n'.join(exec_lines)
                    description = self.parameter_descriptions.get('Exec', 'Bloc de commandes à exécuter')
                    
                    file_data.append({
                        'Section': current_section,
                        'Nom_Section': current_section_name,
                        'Paramètre': 'Exec',
                        'Valeur': param_value,
                        'Description': description
                    })
                
                i += 1
                continue
            
            # Paramètres normaux
            param_match = re.match(r'(\w+)\s+(.+)', line)
            if param_match and current_section:
                param_name = param_match.group(1)
                param_value = param_match.group(2).strip()
                
                # Traitement spécial pour les lignes Exec simples (sans balises)
                if param_name.lower() == 'exec':
                    # Collecter toutes les lignes Exec consécutives
                    exec_lines = [param_value]
                    j = i + 1
                    
                    while j < len(lines):
                        next_line = lines[j].strip()
                        if not next_line:
                            j += 1
                            continue
                        
                        # Vérifier si c'est une autre ligne Exec
                        next_exec_match = re.match(r'Exec\s+(.+)', next_line, re.IGNORECASE)
                        if next_exec_match:
                            exec_lines.append(next_exec_match.group(1).strip())
                            j += 1
                        else:
                            break
                    
                    # Joindre toutes les lignes Exec avec des séparateurs
                    param_value = ' | '.join(exec_lines)
                    i = j - 1  # Ajuster l'index pour éviter de retraiter les lignes Exec
                
                # Supprimer les guillemets si présents
                if param_value.startswith('"') and param_value.endswith('"'):
                    param_value = param_value[1:-1]
                elif param_value.startswith("'") and param_value.endswith("'"):
                    param_value = param_value[1:-1]
                
                # Stocker les informations importantes pour la cartographie
                if param_name == 'Module':
                    section_module = param_value
                
                # Traitement spécial pour les routes
                if current_section == 'Route' and param_name == 'Path':
                    priority = section_details.get('Priority')
                    condition = section_details.get('Condition')
                    flow_mapper.add_route(current_section_name, param_value, priority, condition)
                
                # Stocker les détails de la section
                section_details[param_name] = param_value
                
                description = self.parameter_descriptions.get(param_name, 'Paramètre spécifique au module')
                
                file_data.append({
                    'Section': current_section,
                    'Nom_Section': current_section_name,
                    'Paramètre': param_name,
                    'Valeur': param_value,
                    'Description': description
                })
            
            i += 1
        
        # Analyser les flux après avoir parsé tout le fichier
        flow_mapper.analyze_flows()
        
        return file_data, flow_mapper
    
    def process_directory(self, directory_path: str) -> None:
        """Traite tous les fichiers .conf dans le répertoire spécifié"""
        data_dir = Path(directory_path)
        
        if not data_dir.exists():
            print(f"Erreur: Le répertoire '{directory_path}' n'existe pas.")
            return
        
        # Rechercher tous les fichiers .conf dans le répertoire
        config_files = list(data_dir.glob("*.conf"))
        
        if not config_files:
            print(f"Aucun fichier .conf trouvé dans le répertoire '{directory_path}'.")
            return
        
        print(f"Traitement de {len(config_files)} fichier(s) de configuration...")
        
        # Traiter chaque fichier
        for config_file in config_files:
            print(f"  - Analyse de {config_file.name}...")
            file_data, flow_mapper = self.parse_config_file(str(config_file))
            if file_data:
                self.files_data[config_file.stem] = file_data
                self.flow_mappers[config_file.stem] = flow_mapper
                self.config_data.extend(file_data)
        
        print(f"✅ {len(self.files_data)} fichier(s) traité(s) avec succès.")
    
    def process_single_file(self, config_file: str) -> None:
        """Traite un seul fichier de configuration"""
        file_path = Path(config_file)
        file_data, flow_mapper = self.parse_config_file(config_file)
        if file_data:
            self.files_data[file_path.stem] = file_data
            self.flow_mappers[file_path.stem] = flow_mapper
            self.config_data = file_data
    
    def display_flow_mapping(self, file_name: str = None):
        """Affiche la cartographie des flux"""
        if file_name and file_name in self.flow_mappers:
            mappers = {file_name: self.flow_mappers[file_name]}
        else:
            mappers = self.flow_mappers
        
        for fname, mapper in mappers.items():
            print(f"\n{'='*80}")
            print(f"CARTOGRAPHIE DES FLUX - {fname.upper()}")
            print(f"{'='*80}")
            
            summary = mapper.get_flow_summary()
            
            print(f"📊 RÉSUMÉ:")
            print(f"  • Routes: {summary['total_routes']}")
            print(f"  • Sections: {summary['total_sections']}")
            print(f"  • Flux: {summary['total_flows']}")
            print(f"  • Inputs: {summary['input_sections']}")
            print(f"  • Outputs: {summary['output_sections']}")
            print(f"  • Processors: {summary['processor_sections']}")
            print(f"  • Extensions: {summary['extension_sections']}")
            print(f"  • Sections non connectées: {summary['unconnected_sections']}")
            
            # Affichage des flux
            if mapper.flows:
                print(f"\n🔄 FLUX DE DONNÉES:")
                flow_headers = ['Route', 'Source', 'Type Source', 'Module Source', '', 'Destination', 'Type Dest', 'Module Dest', 'Priorité', 'Condition']
                flow_data = mapper.get_flow_table_data()
                
                if TABULATE_AVAILABLE:
                    print(tabulate(flow_data, headers=flow_headers, tablefmt='grid'))
                else:
                    self._display_simple_table(flow_data, flow_headers)
            
            # Affichage des sections
            print(f"\n📋 SECTIONS:")
            section_headers = ['Nom Section', 'Type', 'Module', 'Routes Connectées', 'Nb Routes']
            section_data = mapper.get_sections_table_data()
            
            if TABULATE_AVAILABLE:
                print(tabulate(section_data, headers=section_headers, tablefmt='grid'))
            else:
                self._display_simple_table(section_data, section_headers)
    
    def _display_simple_table(self, data, headers):
        """Affiche un tableau simple sans tabulate"""
        if not data:
            return
        
        # Calculer les largeurs des colonnes
        col_widths = [len(h) for h in headers]
        
        for row in data:
            for i, val in enumerate(row):
                if i < len(col_widths):
                    col_widths[i] = max(col_widths[i], len(str(val)))
        
        # Afficher le tableau
        def print_separator():
            print('+' + '+'.join('-' * (w + 2) for w in col_widths) + '+')
        
        def print_row(values):
            row = '|'
            for i, val in enumerate(values):
                if i < len(col_widths):
                    row += f' {str(val).ljust(col_widths[i])} |'
            print(row)
        
        print_separator()
        print_row(headers)
        print_separator()
        
        for row in data:
            print_row(row)
        
        print_separator()
    
    def save_flow_mapping_to_csv(self, output_dir: str = "output") -> None:
        """Sauvegarde la cartographie des flux en CSV"""
        import csv
        
        if not self.flow_mappers:
            print("Aucune cartographie de flux à sauvegarder.")
            return
        
        # Créer le répertoire de sortie s'il n'existe pas
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        for filename, mapper in self.flow_mappers.items():
            # Fichier pour les flux
            flows_filename = output_path / f"{filename}_flows.csv"
            flow_headers = ['Route', 'Source', 'Type_Source', 'Module_Source', 'Destination', 'Type_Destination', 'Module_Destination', 'Priorité', 'Condition']
            
            try:
                with open(flows_filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    writer.writerow(flow_headers)
                    
                    for flow in mapper.flows:
                        writer.writerow([
                            flow['route'],
                            flow['source'],
                            flow['source_type'],
                            flow['source_module'],
                            flow['target'],
                            flow['target_type'],
                            flow['target_module'],
                            flow['priority'] or 'N/A',
                            flow['condition'] or 'N/A'
                        ])
                
                print(f"✅ Cartographie des flux générée: {flows_filename}")
                
            except Exception as e:
                print(f"❌ Erreur lors de la création du fichier de flux {flows_filename}: {e}")
            
            # Fichier pour les sections
            sections_filename = output_path / f"{filename}_sections.csv"
            section_headers = ['Nom_Section', 'Type', 'Module', 'Routes_Connectées', 'Nb_Routes']
            
            try:
                with open(sections_filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    writer.writerow(section_headers)
                    
                    for section_name, section_info in mapper.sections.items():
                        connected_routes = ', '.join(section_info['connected_routes']) if section_info['connected_routes'] else 'Non connecté'
                        
                        writer.writerow([
                            section_name,
                            section_info['type'],
                            section_info['module'] or 'N/A',
                            connected_routes,
                            len(section_info['connected_routes'])
                        ])
                
                print(f"✅ Cartographie des sections générée: {sections_filename}")
                
            except Exception as e:
                print(f"❌ Erreur lors de la création du fichier de sections {sections_filename}: {e}")
    
    def display_results_simple_table(self, data: List[Dict]) -> None:
        """Affiche les résultats sous forme de tableau simple sans tabulate"""
        if not data:
            print("Aucun paramètre trouvé dans le fichier de configuration.")
            return
        
        # Calculer les largeurs des colonnes
        headers = ['Section', 'Nom Section', 'Paramètre', 'Valeur', 'Description']
        col_widths = [len(h) for h in headers]
        
        for item in data:
            values = [
                item['Section'],
                item['Nom_Section'],
                item['Paramètre'],
                item['Valeur'][:80] + '...' if len(item['Valeur']) > 80 else item['Valeur'],
                item['Description'][:80] + '...' if len(item['Description']) > 80 else item['Description']
            ]
            for i, val in enumerate(values):
                col_widths[i] = max(col_widths[i], len(str(val)))
        
        # Afficher le tableau
        def print_separator():
            print('+' + '+'.join('-' * (w + 2) for w in col_widths) + '+')
        
        def print_row(values):
            row = '|'
            for i, val in enumerate(values):
                row += f' {str(val).ljust(col_widths[i])} |'
            print(row)
        
        print_separator()
        print_row(headers)
        print_separator()
        
        for item in data:
            values = [
                item['Section'],
                item['Nom_Section'],
                item['Paramètre'],
                item['Valeur'][:80] + '...' if len(item['Valeur']) > 80 else item['Valeur'],
                item['Description'][:80] + '...' if len(item['Description']) > 80 else item['Description']
            ]
            print_row(values)
        
        print_separator()
    
    def save_to_csv_multiple(self, output_dir: str = "output") -> None:
        """Sauvegarde les résultats dans des fichiers CSV séparés par fichier de configuration"""
        import csv
        
        if not self.files_data:
            print("Aucune donnée à sauvegarder.")
            return
        
        # Créer le répertoire de sortie s'il n'existe pas
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        headers = ['Section', 'Nom_Section', 'Paramètre', 'Valeur', 'Description']
        
        for filename, file_data in self.files_data.items():
            csv_filename = output_path / f"{filename}_analysis.csv"
            
            try:
                with open(csv_filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    writer.writerow(headers)
                    
                    for item in file_data:
                        writer.writerow([
                            item['Section'],
                            item['Nom_Section'],
                            item['Paramètre'],
                            item['Valeur'],
                            item['Description']
                        ])
                
                print(f"✅ Fichier CSV généré: {csv_filename}")
                print(f"📊 {len(file_data)} paramètres exportés pour {filename}")
                
            except Exception as e:
                print(f"❌ Erreur lors de la création du fichier CSV {csv_filename}: {e}")
    
    def save_to_csv(self, filename: str) -> None:
        """Sauvegarde les résultats dans un fichier CSV unique"""
        import csv
        
        if not self.config_data:
            print("Aucune donnée à sauvegarder.")
            return
        
        headers = ['Fichier', 'Section', 'Nom_Section', 'Paramètre', 'Valeur', 'Description']
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(headers)
                
                for file_name, file_data in self.files_data.items():
                    for item in file_data:
                        writer.writerow([
                            file_name,
                            item['Section'],
                            item['Nom_Section'],
                            item['Paramètre'],
                            item['Valeur'],
                            item['Description']
                        ])
            
            print(f"✅ Fichier CSV consolidé généré: {filename}")
            print(f"📊 {len(self.config_data)} paramètres exportés")
            
        except Exception as e:
            print(f"❌ Erreur lors de la création du fichier CSV: {e}")
    
    def save_to_excel(self, filename: str) -> None:
        """Sauvegarde les résultats dans un fichier Excel avec un onglet par fichier de configuration + cartographie des flux"""
        if not OPENPYXL_AVAILABLE:
            print("❌ Module 'openpyxl' requis pour générer des fichiers Excel.")
            print("Installation: python3 -m pip install openpyxl")
            return
        
        if not self.files_data:
            print("Aucune donnée à sauvegarder.")
            return
        
        try:
            # Créer un nouveau classeur
            wb = openpyxl.Workbook()
            
            # Supprimer la feuille par défaut
            wb.remove(wb.active)
            
            # En-têtes
            headers = ['Section', 'Nom Section', 'Paramètre', 'Valeur', 'Description']
            
            # Style pour les en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Style pour les données
            data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            # Couleurs alternées pour les lignes
            light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            # Police monospace pour les paramètres Exec
            monospace_font = Font(name='Consolas', size=9)
            
            # Créer un onglet pour chaque fichier de configuration
            for file_name, file_data in self.files_data.items():
                # Créer une nouvelle feuille avec le nom du fichier
                ws = wb.create_sheet(title=file_name[:31])  # Excel limite les noms d'onglets à 31 caractères
                
                # Ajouter les en-têtes
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Ajouter les données
                for row, item in enumerate(file_data, 2):
                    data = [
                        str(item['Section']),
                        str(item['Nom_Section']),
                        str(item['Paramètre']),
                        str(item['Valeur']),
                        str(item['Description'])
                    ]
                    
                    for col, value in enumerate(data, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        # Couleur alternée pour les lignes paires
                        if row % 2 == 0:
                            cell.fill = light_fill
                        
                        # Style spécial pour les paramètres Exec (colonne Valeur)
                        if col == 4 and str(item['Paramètre']).lower() == 'exec':
                            cell.font = monospace_font
                
                # Ajuster la largeur des colonnes
                column_widths = [12, 15, 15, 50, 40]  # Section, Nom Section, Paramètre, Valeur, Description
                
                for col_num, width in enumerate(column_widths, 1):
                    col_letter = get_column_letter(col_num)
                    ws.column_dimensions[col_letter].width = width
                
                # Ajuster la hauteur des lignes pour les paramètres Exec
                for row in range(2, len(file_data) + 2):
                    cell_value = ws.cell(row=row, column=4).value  # Colonne Valeur
                    if cell_value and ('\n' in str(cell_value) or '|' in str(cell_value)):  # Paramètres Exec multi-lignes
                        ws.row_dimensions[row].height = 60.0
                    else:
                        ws.row_dimensions[row].height = 20.0
                
                # Figer la première ligne
                ws.freeze_panes = "A2"
            
            # Ajouter les onglets de cartographie des flux
            for file_name, mapper in self.flow_mappers.items():
                # Onglet pour les flux
                flows_ws = wb.create_sheet(title=f"{file_name[:25]}_Flux")
                
                # En-têtes pour les flux
                flow_headers = ['Route', 'Source', 'Type Source', 'Module Source', 'Destination', 'Type Dest', 'Module Dest', 'Priorité', 'Condition']
                
                for col, header in enumerate(flow_headers, 1):
                    cell = flows_ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Données des flux
                for row, flow in enumerate(mapper.flows, 2):
                    flow_data = [
                        flow['route'],
                        flow['source'],
                        flow['source_type'],
                        flow['source_module'],
                        flow['target'],
                        flow['target_type'],
                        flow['target_module'],
                        flow['priority'] or 'N/A',
                        flow['condition'] or 'N/A'
                    ]
                    
                    for col, value in enumerate(flow_data, 1):
                        cell = flows_ws.cell(row=row, column=col, value=str(value))
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        if row % 2 == 0:
                            cell.fill = light_fill
                
                # Ajuster les colonnes pour les flux
                flow_column_widths = [12, 15, 12, 15, 15, 12, 15, 10, 20]
                for col_num, width in enumerate(flow_column_widths, 1):
                    col_letter = get_column_letter(col_num)
                    flows_ws.column_dimensions[col_letter].width = width
                
                flows_ws.freeze_panes = "A2"
                
                # Onglet pour les sections
                sections_ws = wb.create_sheet(title=f"{file_name[:23]}_Sections")
                
                # En-têtes pour les sections
                section_headers = ['Nom Section', 'Type', 'Module', 'Routes Connectées', 'Nb Routes']
                
                for col, header in enumerate(section_headers, 1):
                    cell = sections_ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Données des sections
                for row, (section_name, section_info) in enumerate(mapper.sections.items(), 2):
                    connected_routes = ', '.join(section_info['connected_routes']) if section_info['connected_routes'] else 'Non connecté'
                    
                    section_data = [
                        section_name,
                        section_info['type'],
                        section_info['module'] or 'N/A',
                        connected_routes,
                        len(section_info['connected_routes'])
                    ]
                    
                    for col, value in enumerate(section_data, 1):
                        cell = sections_ws.cell(row=row, column=col, value=str(value))
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        if row % 2 == 0:
                            cell.fill = light_fill
                
                # Ajuster les colonnes pour les sections
                section_column_widths = [20, 12, 15, 40, 10]
                for col_num, width in enumerate(section_column_widths, 1):
                    col_letter = get_column_letter(col_num)
                    sections_ws.column_dimensions[col_letter].width = width
                
                sections_ws.freeze_panes = "A2"
            
            # Ajouter une feuille de statistiques consolidées
            stats_ws = wb.create_sheet("Statistiques")
            stats = self.get_consolidated_statistics()
            
            # En-tête de la feuille statistiques
            stats_ws.cell(row=1, column=1, value="Statistiques Consolidées - Configuration NXLog")
            stats_ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
            stats_ws.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Données statistiques consolidées
            row_num = 3
            stats_ws.cell(row=row_num, column=1, value="STATISTIQUES GLOBALES").font = Font(bold=True, size=12)
            row_num += 1
            
            global_stats = [
                ("Nombre total de fichiers traités", str(len(self.files_data))),
                ("Nombre total de paramètres", str(stats['total_parameters'])),
                ("Nombre total de sections", str(stats['total_sections'])),
                ("Nombre total de modules", str(stats['total_modules'])),
                ("Nombre total de paramètres Exec", str(stats['total_exec_count'])),
                ("", ""),
                ("Sections trouvées", ", ".join(stats['all_sections'])),
                ("Modules utilisés", ", ".join(stats['all_modules']))
            ]
            
            for label, value in global_stats:
                stats_ws.cell(row=row_num, column=1, value=str(label)).font = Font(bold=True)
                stats_ws.cell(row=row_num, column=2, value=str(value))
                row_num += 1
            
            # Statistiques de cartographie des flux
            row_num += 2
            stats_ws.cell(row=row_num, column=1, value="STATISTIQUES DES FLUX").font = Font(bold=True, size=12)
            row_num += 1
            
            total_routes = sum(len(mapper.routes) for mapper in self.flow_mappers.values())
            total_flows = sum(len(mapper.flows) for mapper in self.flow_mappers.values())
            
            flow_stats = [
                ("Nombre total de routes", str(total_routes)),
                ("Nombre total de flux", str(total_flows))
            ]
            
            for label, value in flow_stats:
                stats_ws.cell(row=row_num, column=1, value=str(label)).font = Font(bold=True)
                stats_ws.cell(row=row_num, column=2, value=str(value))
                row_num += 1
            
            # Statistiques par fichier
            row_num += 2
            stats_ws.cell(row=row_num, column=1, value="STATISTIQUES PAR FICHIER").font = Font(bold=True, size=12)
            row_num += 1
            
            for file_name, file_stats in stats['files_stats'].items():
                stats_ws.cell(row=row_num, column=1, value=f"Fichier: {file_name}").font = Font(bold=True)
                row_num += 1
                
                mapper = self.flow_mappers.get(file_name)
                flow_summary = mapper.get_flow_summary() if mapper else {}
                
                file_data = [
                    ("  Paramètres", str(file_stats['parameters'])),
                    ("  Sections", str(file_stats['sections'])),
                    ("  Modules", str(file_stats['modules'])),
                    ("  Paramètres Exec", str(file_stats['exec_count'])),
                    ("  Routes", str(flow_summary.get('total_routes', 0))),
                    ("  Flux", str(flow_summary.get('total_flows', 0)))
                ]
                
                for label, value in file_data:
                    stats_ws.cell(row=row_num, column=1, value=str(label))
                    stats_ws.cell(row=row_num, column=2, value=str(value))
                    row_num += 1
                
                row_num += 1
            
            # Ajuster les colonnes de la feuille statistiques
            stats_ws.column_dimensions['A'].width = 30.0
            stats_ws.column_dimensions['B'].width = 50.0
            
            # Sauvegarder le fichier
            wb.save(filename)
            
            print(f"✅ Fichier Excel généré: {filename}")
            print(f"📊 {len(self.files_data)} onglet(s) de configuration créé(s)")
            print(f"🔄 {len(self.flow_mappers)} onglet(s) de cartographie des flux créé(s)")
            print(f"📋 {len(self.config_data)} paramètres exportés au total")
            
        except Exception as e:
            print(f"❌ Erreur lors de la création du fichier Excel: {e}")
            import traceback
            traceback.print_exc()
    
    def display_results(self, output_format: str = 'table') -> None:
        """Affiche les résultats sous forme de tableau"""
        if not self.config_data:
            print("Aucun paramètre trouvé dans les fichiers de configuration.")
            return
        
        headers = ['Fichier', 'Section', 'Nom Section', 'Paramètre', 'Valeur', 'Description']
        
        # Préparer les données pour l'affichage
        table_data = []
        for file_name, file_data in self.files_data.items():
            for item in file_data:
                # Pour les paramètres Exec, afficher plus de caractères
                max_val_length = 100 if item['Paramètre'].lower() == 'exec' else 50
                display_value = item['Valeur']
                
                # Pour les blocs Exec multi-lignes, remplacer les retours à la ligne par des séparateurs
                if item['Paramètre'].lower() == 'exec' and '\n' in display_value:
                    display_value = display_value.replace('\n', ' | ')
                
                table_data.append([
                    file_name,
                    item['Section'],
                    item['Nom_Section'],
                    item['Paramètre'],
                    display_value[:max_val_length] + '...' if len(display_value) > max_val_length else display_value,
                    item['Description'][:80] + '...' if len(item['Description']) > 80 else item['Description']
                ])
        
        if output_format == 'table':
            if TABULATE_AVAILABLE:
                print(tabulate(table_data, headers=headers, tablefmt='grid'))
            else:
                print("Module 'tabulate' non disponible. Utilisation du format simple:")
                print()
                self.display_results_simple_table(self.config_data)
        elif output_format == 'csv':
            import csv
            import io
            output = io.StringIO()
            writer = csv.writer(output, delimiter=';')
            writer.writerow(headers)
            writer.writerows(table_data)
            print(output.getvalue())
        elif output_format == 'json':
            import json
            json_data = {}
            for file_name, file_data in self.files_data.items():
                json_data[file_name] = file_data
            print(json.dumps(json_data, indent=2, ensure_ascii=False))
    
    def get_statistics(self, data: List[Dict]) -> Dict:
        """Retourne des statistiques sur une configuration"""
        stats = {}
        sections = set()
        modules = set()
        exec_count = 0
        
        for item in data:
            sections.add(item['Section'])
            if item['Paramètre'] == 'Module':
                modules.add(item['Valeur'])
            if item['Paramètre'].lower() == 'exec':
                exec_count += 1
        
        stats['total_parameters'] = len(data)
        stats['sections_count'] = len(sections)
        stats['modules_count'] = len(modules)
        stats['exec_count'] = exec_count
        stats['sections'] = list(sections)
        stats['modules'] = list(modules)
        
        return stats
    
    def get_consolidated_statistics(self) -> Dict:
        """Retourne des statistiques consolidées pour tous les fichiers"""
        consolidated_stats = {
            'total_parameters': 0,
            'total_sections': 0,
            'total_modules': 0,
            'total_exec_count': 0,
            'all_sections': set(),
            'all_modules': set(),
            'files_stats': {}
        }
        
        for file_name, file_data in self.files_data.items():
            file_stats = self.get_statistics(file_data)
            
            consolidated_stats['total_parameters'] += file_stats['total_parameters']
            consolidated_stats['total_exec_count'] += file_stats['exec_count']
            consolidated_stats['all_sections'].update(file_stats['sections'])
            consolidated_stats['all_modules'].update(file_stats['modules'])
            
            consolidated_stats['files_stats'][file_name] = {
                'parameters': file_stats['total_parameters'],
                'sections': file_stats['sections_count'],
                'modules': file_stats['modules_count'],
                'exec_count': file_stats['exec_count']
            }
        
        consolidated_stats['total_sections'] = len(consolidated_stats['all_sections'])
        consolidated_stats['total_modules'] = len(consolidated_stats['all_modules'])
        consolidated_stats['all_sections'] = list(consolidated_stats['all_sections'])
        consolidated_stats['all_modules'] = list(consolidated_stats['all_modules'])
        
        return consolidated_stats
    
    def display_statistics(self) -> None:
        """Affiche les statistiques consolidées"""
        if len(self.files_data) == 1:
            # Affichage pour un seul fichier
            stats = self.get_statistics(self.config_data)
            
            print("\n" + "="*50)
            print("STATISTIQUES DE CONFIGURATION")
            print("="*50)
            print(f"Nombre total de paramètres: {stats['total_parameters']}")
            print(f"Nombre de sections: {stats['sections_count']}")
            print(f"Nombre de modules: {stats['modules_count']}")
            print(f"Nombre de paramètres Exec: {stats['exec_count']}")
            
            if stats['sections']:
                print(f"\nSections trouvées: {', '.join(stats['sections'])}")
            
            if stats['modules']:
                print(f"Modules utilisés: {', '.join(stats['modules'])}")
            
            # Afficher les statistiques de flux pour un seul fichier
            if self.flow_mappers:
                file_name = list(self.flow_mappers.keys())[0]
                mapper = self.flow_mappers[file_name]
                flow_summary = mapper.get_flow_summary()
                
                print(f"\nStatistiques des flux:")
                print(f"Routes: {flow_summary['total_routes']}")
                print(f"Flux: {flow_summary['total_flows']}")
                print(f"Sections non connectées: {flow_summary['unconnected_sections']}")
            
            print("="*50)
        else:
            # Affichage consolidé pour plusieurs fichiers
            stats = self.get_consolidated_statistics()
            
            print("\n" + "="*60)
            print("STATISTIQUES CONSOLIDÉES DE CONFIGURATION")
            print("="*60)
            print(f"Nombre de fichiers traités: {len(self.files_data)}")
            print(f"Nombre total de paramètres: {stats['total_parameters']}")
            print(f"Nombre total de sections: {stats['total_sections']}")
            print(f"Nombre total de modules: {stats['total_modules']}")
            print(f"Nombre total de paramètres Exec: {stats['total_exec_count']}")
            
            # Statistiques consolidées des flux
            total_routes = sum(len(mapper.routes) for mapper in self.flow_mappers.values())
            total_flows = sum(len(mapper.flows) for mapper in self.flow_mappers.values())
            
            print(f"Nombre total de routes: {total_routes}")
            print(f"Nombre total de flux: {total_flows}")
            
            if stats['all_sections']:
                print(f"\nSections trouvées: {', '.join(stats['all_sections'])}")
            
            if stats['all_modules']:
                print(f"Modules utilisés: {', '.join(stats['all_modules'])}")
            
            print("\n" + "-"*60)
            print("DÉTAIL PAR FICHIER")
            print("-"*60)
            
            for file_name, file_stats in stats['files_stats'].items():
                mapper = self.flow_mappers.get(file_name)
                flow_summary = mapper.get_flow_summary() if mapper else {}
                
                print(f"\n📄 {file_name}:")
                print(f"  • Paramètres: {file_stats['parameters']}")
                print(f"  • Sections: {file_stats['sections']}")
                print(f"  • Modules: {file_stats['modules']}")
                print(f"  • Paramètres Exec: {file_stats['exec_count']}")
                print(f"  • Routes: {flow_summary.get('total_routes', 0)}")
                print(f"  • Flux: {flow_summary.get('total_flows', 0)}")
            
            print("="*60)

def parse_nxlog_config(config_file):
    """Parse un fichier de configuration NXLog et retourne les données structurées"""
    import json
    import csv
    from pathlib import Path
    import re
    
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            content = f.read()
    except FileNotFoundError:
        print(f"Erreur: Fichier '{config_file}' non trouvé")
        return None
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier: {e}")
        return None
    
    # Supprimer les commentaires
    content = re.sub(r'#.*$', '', content, flags=re.MULTILINE)
    
    # Initialiser les structures de données
    config_data = []
    flow_mapper = FlowMapper()
    
    current_section = None
    current_section_name = None
    lines = content.split('\n')
    i = 0
    
    # Variables pour stocker les informations de section
    section_module = None
    section_details = {}
    
    while i < len(lines):
        line = lines[i].strip()
        if not line:
            i += 1
            continue
            
        # Début de section
        section_match = re.match(r'<(\w+)\s+([^>]+)>', line, re.IGNORECASE)
        if section_match:
            current_section = section_match.group(1)
            current_section_name = section_match.group(2)
            section_module = None
            section_details = {}
            i += 1
            continue
        
        # Fin de section
        if re.match(r'</\w+>', line, re.IGNORECASE):
            # Ajouter la section au flow mapper
            if current_section and current_section_name:
                flow_mapper.add_section(current_section, current_section_name, section_module)
            
            current_section = None
            current_section_name = None
            section_module = None
            section_details = {}
            i += 1
            continue
        
        # Paramètres
        param_match = re.match(r'(\w+)\s+(.+)', line)
        if param_match and current_section:
            param_name = param_match.group(1)
            param_value = param_match.group(2).strip()
            
            # Supprimer les guillemets si présents
            if param_value.startswith('"') and param_value.endswith('"'):
                param_value = param_value[1:-1]
            elif param_value.startswith("'") and param_value.endswith("'"):
                param_value = param_value[1:-1]
            
            # Stocker les informations importantes
            if param_name == 'Module':
                section_module = param_value
            
            # Traitement spécial pour les routes
            if current_section == 'Route' and param_name == 'Path':
                priority = section_details.get('Priority')
                condition = section_details.get('Condition')
                flow_mapper.add_route(current_section_name, param_value, priority, condition)
            
            section_details[param_name] = param_value
            
            config_data.append({
                'section_type': current_section,
                'section_name': current_section_name,
                'parameter': param_name,
                'value': param_value
            })
        
        i += 1
    
    # Analyser les flux
    flow_data = flow_mapper.analyze_flows()
    
    return {
        'config': config_data,
        'flows': flow_data
    }

def save_to_excel(all_data, filename):
    """Sauvegarde les données dans un fichier Excel avec onglets séparés"""
    if not OPENPYXL_AVAILABLE:
        print("❌ Module 'openpyxl' requis pour générer des fichiers Excel.")
        print("Installation: python3 -m pip install openpyxl")
        return
    
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Supprimer la feuille par défaut
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Style pour les données
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        
        for file_name, data in all_data.items():
            if not data:
                continue
                
            config_data = data.get('config', [])
            flow_data = data.get('flows', {})
            
            # Onglet pour la configuration
            ws_config = wb.create_sheet(title=f"{file_name[:25]}_Config")
            
            # En-têtes configuration
            config_headers = ['Type Section', 'Nom Section', 'Paramètre', 'Valeur']
            for col, header in enumerate(config_headers, 1):
                cell = ws_config.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Données configuration
            for row, item in enumerate(config_data, 2):
                row_data = [
                    item['section_type'],
                    item['section_name'],
                    item['parameter'],
                    str(item['value'])
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws_config.cell(row=row, column=col, value=value)
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    if row % 2 == 0:
                        cell.fill = light_fill
            
            # Ajuster les colonnes
            config_column_widths = [15, 20, 20, 50]
            for col_num, width in enumerate(config_column_widths, 1):
                col_letter = get_column_letter(col_num)
                ws_config.column_dimensions[col_letter].width = width
            
            ws_config.freeze_panes = "A2"
            
            # Onglet pour les flux si disponibles
            if flow_data and flow_data.get('flows'):
                ws_flows = wb.create_sheet(title=f"{file_name[:25]}_Flows")
                
                # En-têtes flux
                flow_headers = ['Route', 'Source', 'Destination', 'Priorité', 'Condition']
                for col, header in enumerate(flow_headers, 1):
                    cell = ws_flows.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Données flux
                for row, flow in enumerate(flow_data['flows'], 2):
                    flow_row_data = [
                        flow['route'],
                        flow['source'],
                        flow['destination'],
                        flow['priority'],
                        flow['condition']
                    ]
                    
                    for col, value in enumerate(flow_row_data, 1):
                        cell = ws_flows.cell(row=row, column=col, value=str(value))
                        cell.alignment = data_alignment
                        cell.border = border
                        
                        if row % 2 == 0:
                            cell.fill = light_fill
                
                # Ajuster les colonnes flux
                flow_column_widths = [15, 20, 20, 10, 30]
                for col_num, width in enumerate(flow_column_widths, 1):
                    col_letter = get_column_letter(col_num)
                    ws_flows.column_dimensions[col_letter].width = width
                
                ws_flows.freeze_panes = "A2"
        
        wb.save(filename)
        print(f"✅ Fichier Excel généré: {filename}")
        
    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier Excel: {e}")

def save_flows_to_csv(flow_data, base_name, output_dir):
    """Sauvegarde les flux dans des fichiers CSV"""
    if not flow_data or not flow_data.get('flows'):
        return
    
    import csv
    
    # Créer le répertoire de sortie s'il n'existe pas
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # Fichier pour les flux
    flows_file = Path(output_dir) / f"{base_name}_flows.csv"
    
    try:
        with open(flows_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['Route', 'Source', 'Destination', 'Priorité', 'Condition'])
            
            for flow in flow_data['flows']:
                writer.writerow([
                    flow['route'],
                    flow['source'],
                    flow['destination'],
                    flow['priority'],
                    flow['condition']
                ])
        
        print(f"✅ Fichier CSV des flux sauvegardé: {flows_file}")
        
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde des flux CSV: {e}")
    
    # Fichier pour les sections
    sections_file = Path(output_dir) / f"{base_name}_sections.csv"
    
    try:
        with open(sections_file, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(['Nom_Section', 'Type', 'Module', 'Connecté'])
            
            sections = flow_data.get('sections', {})
            unconnected = set(flow_data.get('unconnected', []))
            
            for section_name, section_info in sections.items():
                writer.writerow([
                    section_name,
                    section_info['type'],
                    section_info['module'],
                    'Non' if section_name in unconnected else 'Oui'
                ])
        
        print(f"✅ Fichier CSV des sections sauvegardé: {sections_file}")
        
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde des sections CSV: {e}")

def generate_synthesis_image_script(script_file):
    """Génère le script pour créer les images de la cartographie de synthèse"""
    script_content = '''#!/bin/bash

# Script de génération d'images pour la cartographie de synthèse NXLog
# Généré automatiquement par nxlog_analyzer.py

echo "🎨 Génération des images de synthèse NXLog..."

# Vérifier si Graphviz est installé
if ! command -v dot &> /dev/null; then
    echo "❌ Erreur: Graphviz n'est pas installé"
    echo "💡 Installation:"
    echo "   Ubuntu/Debian: sudo apt-get install graphviz"
    echo "   CentOS/RHEL: sudo yum install graphviz"
    echo "   macOS: brew install graphviz"
    exit 1
fi

# Générer les différents formats
echo "📊 Génération PNG..."
dot -Tpng nxlog_synthesis_flow.dot -o nxlog_synthesis_flow.png

echo "🎯 Génération SVG..."
dot -Tsvg nxlog_synthesis_flow.dot -o nxlog_synthesis_flow.svg

echo "📄 Génération PDF..."
dot -Tpdf nxlog_synthesis_flow.dot -o nxlog_synthesis_flow.pdf

echo "✅ Images de synthèse générées avec succès!"
echo "📁 Fichiers créés:"
echo "   • nxlog_synthesis_flow.png (image bitmap)"
echo "   • nxlog_synthesis_flow.svg (image vectorielle)"
echo "   • nxlog_synthesis_flow.pdf (document imprimable)"
'''
    
    with open(script_file, 'w', encoding='utf-8') as f:
        f.write(script_content)
    
    # Rendre le script exécutable
    os.chmod(script_file, 0o755)

def save_graphviz(flow_data, base_name, output_dir):
    """Sauvegarde les fichiers Graphviz (.dot)"""
    if not flow_data or 'graphviz' not in flow_data:
        return
    
    # Créer le répertoire de sortie s'il n'existe pas
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # Sauvegarder le fichier .dot
    dot_file = Path(output_dir) / f"{base_name}_flow.dot"
    
    try:
        with open(dot_file, 'w', encoding='utf-8') as f:
            f.write(flow_data['graphviz'])
        
        print(f"✅ Fichier Graphviz sauvegardé: {dot_file}")
        
        # Créer aussi un fichier de commandes pour générer les images
        commands_file = Path(output_dir) / f"{base_name}_generate_images.sh"
        with open(commands_file, 'w', encoding='utf-8') as f:
            f.write("#!/bin/bash\n")
            f.write("# Script pour générer les images à partir du fichier .dot\n")
            f.write("# Assurez-vous d'avoir Graphviz installé: sudo apt-get install graphviz\n\n")
            f.write(f"# Générer une image PNG\n")
            f.write(f"dot -Tpng {base_name}_flow.dot -o {base_name}_flow.png\n\n")
            f.write(f"# Générer une image SVG (vectorielle)\n")
            f.write(f"dot -Tsvg {base_name}_flow.dot -o {base_name}_flow.svg\n\n")
            f.write(f"# Générer un PDF\n")
            f.write(f"dot -Tpdf {base_name}_flow.dot -o {base_name}_flow.pdf\n\n")
            f.write(f"echo 'Images générées avec succès !'\n")
        
        # Rendre le script exécutable
        import stat
        commands_file.chmod(commands_file.stat().st_mode | stat.S_IEXEC)
        
        print(f"✅ Script de génération d'images créé: {commands_file}")
        print("💡 Pour générer les images, exécutez:")
        print(f"   cd {output_dir} && ./{base_name}_generate_images.sh")
        
    except Exception as e:
        print(f"❌ Erreur lors de la sauvegarde Graphviz: {e}")

def display_global_stats(all_data):
    """Affiche les statistiques globales pour tous les fichiers"""
    total_sections = 0
    total_parameters = 0
    total_flows = 0
    all_section_types = set()
    all_modules = set()
    
    print("\n" + "="*60)
    print("STATISTIQUES GLOBALES")
    print("="*60)
    
    for file_name, data in all_data.items():
        if not data:
            continue
            
        config_data = data.get('config', [])
        flow_data = data.get('flows', {})
        
        file_sections = set()
        file_modules = set()
        
        for item in config_data:
            file_sections.add(f"{item['section_type']}:{item['section_name']}")
            all_section_types.add(item['section_type'])
            
            if item['parameter'] == 'Module':
                file_modules.add(item['value'])
                all_modules.add(item['value'])
        
        file_flows = len(flow_data.get('flows', []))
        
        print(f"\n📄 {file_name}:")
        print(f"  • Sections: {len(file_sections)}")
        print(f"  • Paramètres: {len(config_data)}")
        print(f"  • Flux: {file_flows}")
        print(f"  • Modules: {', '.join(file_modules) if file_modules else 'Aucun'}")
        
        total_sections += len(file_sections)
        total_parameters += len(config_data)
        total_flows += file_flows
    
    print(f"\n📊 TOTAUX:")
    print(f"  • Fichiers traités: {len(all_data)}")
    print(f"  • Sections totales: {total_sections}")
    print(f"  • Paramètres totaux: {total_parameters}")
    print(f"  • Flux totaux: {total_flows}")
    print(f"  • Types de sections: {', '.join(sorted(all_section_types))}")
    print(f"  • Modules utilisés: {', '.join(sorted(all_modules))}")
    
    print("="*60)

def display_flows(flow_data, file_name):
    """Affiche la cartographie des flux pour un fichier"""
    if not flow_data or not flow_data.get('flows'):
        print(f"Aucun flux trouvé pour {file_name}")
        return
    
    print(f"\n🔄 CARTOGRAPHIE DES FLUX - {file_name.upper()}")
    print("="*60)
    
    stats = flow_data.get('stats', {})
    print(f"📊 Statistiques:")
    print(f"  • Routes: {stats.get('total_routes', 0)}")
    print(f"  • Sections: {stats.get('total_sections', 0)}")
    print(f"  • Flux: {stats.get('total_flows', 0)}")
    print(f"  • Sections non connectées: {stats.get('unconnected_sections', 0)}")
    
    # Afficher les flux
    flows = flow_data['flows']
    if flows:
        print(f"\n📋 Détail des flux:")
        
        if TABULATE_AVAILABLE:
            flow_table = []
            for flow in flows:
                flow_table.append([
                    flow['route'],
                    flow['source'],
                    flow['destination'],
                    flow['priority'],
                    flow['condition'][:50] + '...' if len(flow['condition']) > 50 else flow['condition']
                ])
            
            headers = ['Route', 'Source', 'Destination', 'Priorité', 'Condition']
            print(tabulate(flow_table, headers=headers, tablefmt='grid'))
        else:
            for flow in flows:
                print(f"  Route: {flow['route']}")
                print(f"    {flow['source']} → {flow['destination']}")
                if flow['priority'] != 'N/A':
                    print(f"    Priorité: {flow['priority']}")
                if flow['condition'] != 'N/A':
                    print(f"    Condition: {flow['condition']}")
                print()
    
    # Afficher les sections non connectées
    unconnected = flow_data.get('unconnected', [])
    if unconnected:
        print(f"⚠️  Sections non connectées: {', '.join(unconnected)}")

def create_sample_config():
    """Crée un exemple de fichier de configuration nxlog"""
    sample_config = """# Configuration nxlog example
# Global settings
LogLevel INFO
LogFile /var/log/nxlog.log
LogToConsole FALSE

# Input module - Windows Event Log
<Input eventlog>
    Module im_msvistalog
    Query <QueryList>\
            <Query Id="0">\
                <Select Path="Application">*</Select>\
            </Query>\
        </QueryList>
    Exec $EventTime = parsedate($EventTimeStr);
    Exec $Hostname = hostname_fqdn();
    Exec if $EventID == 1000 drop();
    
    <Exec>
        # Bloc Exec multi-lignes
        $EventTime = parsedate($EventTimeStr);
        $Hostname = hostname_fqdn();
        if $EventID == 1000 drop();
        $Message = "Processed: " + $raw_event;
    </Exec>
</Input>

# Input module - File
<Input file>
    Module im_file
    File "/var/log/app/*.log"
    InputType LineBased
    SavePos TRUE
    ReadFromLast TRUE
    PollInterval 1
    Exec $Message = $raw_event;
    Exec $SourceName = "MyApp";
    Exec if $Message =~ /DEBUG/ drop();
    Exec $Severity = INFO;
    
    <Exec>
        # Traitement complexe des logs
        if $Message =~ /ERROR/
        {
            $Severity = ERROR;
            $Priority = HIGH;
        }
        else if $Message =~ /WARN/
        {
            $Severity = WARNING;
            $Priority = MEDIUM;
        }
        else
        {
            $Severity = INFO;
            $Priority = LOW;
        }
        
        # Ajout de métadonnées
        $ProcessedTime = now();
        $Source = "Application";
    </Exec>
</Input>

# Processor for pattern matching
<Processor pattern>
    Module pm_pattern
    PatternFile /etc/nxlog/patterns.xml
    Exec $ParsedMessage = $PatternName + ": " + $Message;
    Exec if not defined($PatternName) $ParsedMessage = "UNKNOWN: " + $Message;
    
    <Exec>
        # Traitement avancé des patterns
        if defined($PatternName)
        {
            $ParsedMessage = $PatternName + ": " + $Message;
            $MatchFound = TRUE;
        }
        else
        {
            $ParsedMessage = "UNKNOWN: " + $Message;
            $MatchFound = FALSE;
        }
        
        # Log des patterns non reconnus
        if not $MatchFound
        {
            log_warning("Pattern non reconnu: " + $Message);
        }
    </Exec>
</Processor>

# Output module - Syslog
<Output syslog>
    Module om_udp
    Host 192.168.1.100
    Port 514
    Facility LOCAL0
    Severity INFO
    OutputType Syslog_RFC3164
    Exec $SyslogFacilityValue = facility_value("local0");
    
    <Exec>
        # Formatage pour syslog
        $SyslogFacilityValue = facility_value("local0");
        $SyslogSeverityValue = severity_value($Severity);
        $raw_event = to_syslog_rfc3164();
    </Exec>
</Output>

# Output module - File
<Output fileout>
    Module om_file
    File "/var/log/processed.log"
    CreateDir TRUE
    Sync TRUE
    FlushInterval 1
    Exec $raw_event = to_json();
</Output>

# Route principale
<Route main>
    Path eventlog, file => pattern => syslog, fileout
    Priority 1
</Route>

# Route de secours
<Route backup>
    Path eventlog => fileout
    Priority 2
    Condition $Severity == ERROR
</Route>

# Extension for CSV parsing
<Extension csv>
    Module xm_csv
    Fields $timestamp, $level, $message
    FieldTypes string, string, string
    Delimiter ,
    QuoteChar "
    EscapeChar \\
</Extension>
"""
    
    # Créer le répertoire data s'il n'existe pas
    data_dir = Path("data")
    data_dir.mkdir(exist_ok=True)
    
    # Créer plusieurs fichiers d'exemple
    configs = {
        "nxlog_sample.conf": sample_config,
        "nxlog_web.conf": """# Configuration pour serveur web
LogLevel INFO
LogFile /var/log/nxlog_web.log

<Input apache_access>
    Module im_file
    File "/var/log/apache2/access.log"
    InputType LineBased
    SavePos TRUE
    Exec parse_apache_access_log();
</Input>

<Input apache_error>
    Module im_file
    File "/var/log/apache2/error.log"
    InputType LineBased
    SavePos TRUE
    Exec parse_apache_error_log();
</Input>

<Processor web_filter>
    Module pm_filter
    Condition $Message =~ /404|500|502|503/
    Exec $AlertLevel = "HIGH";
</Processor>

<Output syslog_web>
    Module om_udp
    Host 192.168.1.200
    Port 514
    Facility LOCAL1
</Output>

<Output alert_file>
    Module om_file
    File "/var/log/web_alerts.log"
    CreateDir TRUE
</Output>

<Route web_logs>
    Path apache_access, apache_error => web_filter => syslog_web
    Priority 1
</Route>

<Route web_alerts>
    Path apache_access, apache_error => web_filter => alert_file
    Priority 2
    Condition $AlertLevel == "HIGH"
</Route>
""",
        "nxlog_db.conf": """# Configuration pour base de données
LogLevel DEBUG
LogFile /var/log/nxlog_db.log

<Input mysql_slow>
    Module im_file
    File "/var/log/mysql/slow.log"
    InputType LineBased
    SavePos TRUE
    
    <Exec>
        # Traitement des requêtes lentes MySQL
        if $raw_event =~ /^# Time:/
        {
            $QueryTime = extract_time($raw_event);
        }
        else if $raw_event =~ /^# Query_time:/
        {
            $Duration = extract_duration($raw_event);
            if $Duration > 5.0
            {
                $Severity = WARNING;
            }
        }
    </Exec>
</Input>

<Input mysql_error>
    Module im_file
    File "/var/log/mysql/error.log"
    InputType LineBased
    SavePos TRUE
</Input>

<Processor db_analyzer>
    Module pm_pattern
    PatternFile /etc/nxlog/db_patterns.xml
    
    <Exec>
        # Analyse des performances DB
        if $Duration > 10.0
        {
            $Priority = "CRITICAL";
        }
        else if $Duration > 5.0
        {
            $Priority = "WARNING";
        }
    </Exec>
</Processor>

<Output db_monitoring>
    Module om_tcp
    Host monitoring.example.com
    Port 1514
    OutputType Syslog_RFC5424
</Output>

<Output db_archive>
    Module om_file
    File "/var/log/db_archive.log"
    CreateDir TRUE
</Output>

<Route db_performance>
    Path mysql_slow => db_analyzer => db_monitoring
    Priority 1
</Route>

<Route db_errors>
    Path mysql_error => db_monitoring, db_archive
    Priority 2
</Route>

<Route db_critical>
    Path mysql_slow => db_analyzer => db_monitoring
    Priority 3
    Condition $Priority == "CRITICAL"
</Route>
"""
    }
    
    for filename, content in configs.items():
        file_path = data_dir / filename
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(content)
    
    print(f"✅ {len(configs)} fichiers d'exemple créés dans le répertoire 'data/':")
    for filename in configs.keys():
        print(f"  - data/{filename}")
    print("Ces fichiers contiennent des exemples de routes et de cartographie des flux.")

def main():
    """Fonction principale"""
    parser = argparse.ArgumentParser(
        description='Analyseur de configuration nxlog avec cartographie des flux',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples d'utilisation:

Fichier unique:
  python3 nxlog_analyzer.py nxlog.conf
  python3 nxlog_analyzer.py nxlog.conf --flows --stats

Répertoire de fichiers:
  python3 nxlog_analyzer.py --directory data
  python3 nxlog_analyzer.py --directory data --flows --stats
  python3 nxlog_analyzer.py --directory data --excel output.xlsx
  python3 nxlog_analyzer.py --directory data --csv-multiple --flows-csv

Création d'exemples:
  python3 nxlog_analyzer.py --create-sample

Visualisation Graphviz:
  python3 nxlog_analyzer.py --directory data --graphviz
        """
    )
    
    parser.add_argument('config_file', nargs='?', help='Fichier de configuration nxlog à analyser')
    parser.add_argument('--create-sample', action='store_true', help='Créer un fichier d\'exemple de configuration')
    parser.add_argument('--directory', '-d', type=str,
                       help='Répertoire contenant les fichiers .conf à analyser')
    parser.add_argument('--flows', action='store_true',
                       help='Afficher la cartographie des flux')
    parser.add_argument('--stats', action='store_true',
                       help='Afficher les statistiques')
    parser.add_argument('--excel', type=str,
                       help='Générer un fichier Excel (ex: --excel output.xlsx)')
    parser.add_argument('--csv-multiple', action='store_true',
                       help='Générer des fichiers CSV séparés dans le répertoire output/')
    parser.add_argument('--flows-csv', action='store_true', 
                       help='Créer des fichiers CSV séparés pour les flux (avec --csv-multiple)')
    
    parser.add_argument('--graphviz', action='store_true',
                       help='Générer des fichiers Graphviz (.dot) pour la visualisation des flux')
    
    args = parser.parse_args()
    
    if args.create_sample:
        create_sample_config()
        return
    
    if not args.config_file and not args.directory:
        parser.print_help()
        return
    
    # Collecter tous les fichiers à traiter
    files_to_process = []
    
    if args.directory:
        data_dir = Path(args.directory)
        if not data_dir.exists():
            print(f"Erreur: Le répertoire '{args.directory}' n'existe pas.")
            return
        
        config_files = list(data_dir.glob("*.conf"))
        if not config_files:
            print(f"Aucun fichier .conf trouvé dans le répertoire '{args.directory}'.")
            return
        
        files_to_process = config_files
        print(f"Traitement de {len(config_files)} fichier(s) de configuration...")
    else:
        config_file = Path(args.config_file)
        if not config_file.exists():
            print(f"Erreur: Le fichier '{args.config_file}' n'existe pas.")
            return
        
        files_to_process = [config_file]
    
    # Traiter tous les fichiers
    all_data = {}
    output_dir = "output"
    
    for config_file in files_to_process:
        print(f"  - Analyse de {config_file.name}...")
        
        data = parse_nxlog_config(str(config_file))
        if data:
            base_name = config_file.stem
            all_data[base_name] = data
            
            # Afficher les flux si demandé
            if args.flows:
                display_flows(data['flows'], base_name)
            
            # Sauvegarder les flux CSV si demandé
            if args.flows_csv:
                save_flows_to_csv(data['flows'], base_name, output_dir)
            
            # Sauvegarder les fichiers Graphviz si demandé
            if args.graphviz:
                save_graphviz(data['flows'], base_name, output_dir)
    
    print(f"\n📊 Résumé: {len(all_data)} fichier(s) traité(s)")
    
    # Actions spéciales pour le répertoire
    if all_data:
        if args.graphviz:
            print(f"\n🎨 Génération des diagrammes Graphviz...")
            # Générer les fichiers Graphviz pour chaque fichier
            for config_file, (config_data, flow_data) in all_configs.items():
                if flow_data['routes']:
                    print(f"   • Génération pour: {os.path.basename(config_file)}")
                    generate_graphviz_files(config_file, flow_data, args.output_dir)
            
            # Toujours générer la cartographie de synthèse si on a des données
            print(f"   • Génération de la synthèse globale...")
            if all_configs:
                generate_synthesis_graphviz(all_configs, args.output_dir)
            
            print(f"\n✅ Fichiers Graphviz générés dans le répertoire: {args.output_dir}")
            print("📁 Fichiers générés:")
            for file in os.listdir(args.output_dir):
                if file.endswith(('.dot', '.sh')):
                    print(f"   • {file}")
            print("\n🔧 Pour générer les images:")
            print(f"   cd {args.output_dir}")
            print("   ./nxlog_synthesis_generate_images.sh")
            
            # Vérifier que le script de synthèse existe
            synthesis_script = os.path.join(args.output_dir, "nxlog_synthesis_generate_images.sh")
            if os.path.exists(synthesis_script):
                print(f"✅ Script de synthèse confirmé: {synthesis_script}")
            else:
                print(f"❌ Script de synthèse manquant: {synthesis_script}")
        
        elif args.flows_csv:
            # Afficher les statistiques globales si demandé
            if args.stats:
                display_global_stats(all_data)
            
            # Générer le fichier Excel si demandé
            if args.excel:
                save_to_excel(all_data, args.excel)
            
            # Générer les fichiers CSV multiples si demandé
            if args.csv_multiple:
                save_to_csv_multiple(all_data, output_dir)
    
    print(f"✅ Traitement terminé. {len(all_data)} fichier(s) analysé(s).")
    
    if args.flows_csv or args.csv_multiple or args.graphviz:
        print(f"📁 Fichiers de sortie disponibles dans le répertoire: {output_dir}")

def save_to_csv_multiple(all_data, output_dir):
    """Sauvegarde les données de configuration dans des fichiers CSV séparés"""
    import csv
    
    # Créer le répertoire de sortie s'il n'existe pas
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    for file_name, data in all_data.items():
        if not data or not data.get('config'):
            continue
        
        csv_file = Path(output_dir) / f"{file_name}_config.csv"
        
        try:
            with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                writer.writerow(['Type_Section', 'Nom_Section', 'Paramètre', 'Valeur'])
                
                for item in data['config']:
                    writer.writerow([
                        item['section_type'],
                        item['section_name'],
                        item['parameter'],
                        item['value']
                    ])
            
            print(f"✅ Fichier CSV de configuration sauvegardé: {csv_file}")
            
        except Exception as e:
            print(f"❌ Erreur lors de la sauvegarde CSV: {e}")
    
    print(f"✅ Fichiers CSV des flux sauvegardés dans le répertoire: {output_dir}")

if __name__ == "__main__":
    main()